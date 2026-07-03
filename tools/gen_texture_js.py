#!/usr/bin/env python3
"""
从 ε_噪声表.xlsx 生成 index.html 用的 ε 噪声 JS 数组（DAY_FACTOR / HOUR_TEX）。

ε 模型（见 xlsx「说明」页）：
    ε(d, h) = 波动幅度[类型] × ( 日因子[日序 d] + 0.4 × 时纹理[星期(d)][h] )
    P(d, h) ← P(d, h) × ( 1 + ε(d, h) )

用法：
    pip install openpyxl          # 或 pip install --break-system-packages openpyxl
    python3 tools/gen_texture_js.py            # 打印 JS，手动粘贴回 index.html 的 ε 数据块
    python3 tools/gen_texture_js.py > eps.js   # 或输出到文件

输出（stdout）：
    const DAY_FACTOR=[...365...];              # 365 天日因子，零均值 σ≈1
    const HOUR_TEX=[[..7..], ... ×24];         # HOUR_TEX[hour]=[周一..周日]
波动幅度表打印到 stderr —— 用于填 index.html 里 PRESETS 各类型的 amp 字段。
"""
import sys
import os
import json

try:
    import openpyxl
except ImportError:
    sys.exit("需要 openpyxl：pip install openpyxl（或加 --break-system-packages）")

XLSX = os.path.join(os.path.dirname(__file__), "..", "ε_噪声表.xlsx")


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # 日因子（365 天，每天一个量级偏差）
    dayf = []
    for i, r in enumerate(wb["日因子"].iter_rows(values_only=True)):
        if i == 0 or r[2] is None:
            continue
        dayf.append(round(float(r[2]), 4))

    # 时纹理：HOUR_TEX[hour] = [周一, 周二, ..., 周日]
    ht = []
    for i, r in enumerate(wb["时纹理"].iter_rows(values_only=True)):
        if i == 0 or r[0] is None:
            continue
        ht.append([round(float(x), 4) for x in r[1:8]])

    # 波动幅度（按负载类型）
    amp = {}
    for i, r in enumerate(wb["波动幅度"].iter_rows(values_only=True)):
        if i > 0 and r[0]:
            amp[r[0]] = r[1]

    assert len(dayf) == 365, f"日因子应为 365 个，实际 {len(dayf)}"
    assert len(ht) == 24 and len(ht[0]) == 7, f"时纹理应为 24×7，实际 {len(ht)}×{len(ht[0])}"

    print(f"const DAY_FACTOR={json.dumps(dayf)};")
    print("const HOUR_TEX=[" + ",".join(
        "[" + ",".join(str(x) for x in row) + "]" for row in ht) + "];")

    print("\n// 波动幅度（填入 index.html 中 PRESETS 各类型的 amp 字段）：", file=sys.stderr)
    for k, v in amp.items():
        print(f"//   {k}: {v}", file=sys.stderr)


if __name__ == "__main__":
    main()
